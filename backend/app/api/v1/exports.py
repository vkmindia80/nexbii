from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from typing import Optional
import io
import csv
import json
from datetime import datetime

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.dashboard import Dashboard
from app.models.query import Query
from app.models.datasource import DataSource
from app.services.query_service import QueryService

query_service = QueryService()

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

router = APIRouter()


@router.get("/query/{query_id}/csv")
async def export_query_to_csv(
    query_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export query results to CSV"""
    # Get query
    query = db.query(Query).filter(Query.id == query_id).first()
    if not query:
        raise HTTPException(status_code=404, detail="Query not found")
    
    # Execute query
    try:
        result = await execute_query_internal(query.datasource_id, query.sql_query, db)
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        if result.get('columns'):
            writer.writerow(result['columns'])
        
        # Write data
        if result.get('data'):
            for row in result['data']:
                writer.writerow(row)
        
        output.seek(0)
        
        # Return as streaming response
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=query_{query_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/query/{query_id}/excel")
async def export_query_to_excel(
    query_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export query results to Excel"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=501, detail="Excel export not available. Install openpyxl.")
    
    # Get query
    query = db.query(Query).filter(Query.id == query_id).first()
    if not query:
        raise HTTPException(status_code=404, detail="Query not found")
    
    # Execute query
    try:
        result = await execute_query_internal(query.datasource_id, query.sql_query, db)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Query Results"
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Write headers
        if result.get('columns'):
            for col_idx, col_name in enumerate(result['columns'], start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
        
        # Write data
        if result.get('data'):
            for row_idx, row_data in enumerate(result['data'], start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=query_{query_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/dashboard/{dashboard_id}/json")
async def export_dashboard_to_json(
    dashboard_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export dashboard configuration to JSON"""
    dashboard = db.query(Dashboard).filter(Dashboard.id == dashboard_id).first()
    if not dashboard:
        raise HTTPException(status_code=404, detail="Dashboard not found")
    
    # Create export data
    export_data = {
        "name": dashboard.name,
        "description": dashboard.description,
        "layout": dashboard.layout,
        "widgets": dashboard.widgets,
        "filters": dashboard.filters,
        "exported_at": datetime.utcnow().isoformat(),
        "version": "1.0"
    }
    
    # Convert to JSON
    json_str = json.dumps(export_data, indent=2)
    
    return StreamingResponse(
        io.BytesIO(json_str.encode('utf-8')),
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=dashboard_{dashboard_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        }
    )


@router.post("/dashboard/{dashboard_id}/pdf")
async def export_dashboard_to_pdf(
    dashboard_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export dashboard to PDF (basic version)"""
    if not PDF_AVAILABLE:
        raise HTTPException(status_code=501, detail="PDF export not available. Install reportlab.")
    
    dashboard = db.query(Dashboard).filter(Dashboard.id == dashboard_id).first()
    if not dashboard:
        raise HTTPException(status_code=404, detail="Dashboard not found")
    
    # Create PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph(f"<b>{dashboard.name}</b>", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 0.2*inch))
    
    # Description
    if dashboard.description:
        desc = Paragraph(dashboard.description, styles['Normal'])
        story.append(desc)
        story.append(Spacer(1, 0.3*inch))
    
    # Export info
    export_info = Paragraph(
        f"Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Exported by: {current_user.email}",
        styles['Normal']
    )
    story.append(export_info)
    story.append(Spacer(1, 0.3*inch))
    
    # Widget summary
    widget_count = len(dashboard.widgets) if dashboard.widgets else 0
    summary = Paragraph(f"<b>Dashboard Summary</b><br/>Total Widgets: {widget_count}", styles['Heading2'])
    story.append(summary)
    story.append(Spacer(1, 0.2*inch))
    
    # List widgets
    if dashboard.widgets:
        widget_data = [["Widget", "Type", "Title"]]
        for idx, widget in enumerate(dashboard.widgets, 1):
            widget_data.append([
                str(idx),
                widget.get('type', 'N/A'),
                widget.get('title', 'Untitled')
            ])
        
        table = Table(widget_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=dashboard_{dashboard_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        }
    )
